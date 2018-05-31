from openpyxl.reader.excel import load_workbook

def readxlsxfile(path):
    file=load_workbook(path)
    sheets=file.sheetnames
    dic={}
    for i in range(len(sheets)):
        sheet=file.worksheets[i]
        sheetinfo=[]

        for lineNum in range(1,sheet.max_row+1):
            linelist = []
            for columnNum in  range(1,sheet.max_column+1):
                value=sheet.cell(row=lineNum,column=columnNum).value
                linelist.append(value)

            sheetinfo.append(linelist)

        dic[file.worksheets[i]]=sheetinfo
    return dic



path="D:\\pythonitem\\demo\\Demo7\\1.xlsx"
dic1=readxlsxfile(path)
print(dic1)

