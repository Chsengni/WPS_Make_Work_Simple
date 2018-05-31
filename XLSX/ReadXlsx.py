from openpyxl.reader.excel import load_workbook

def readxlsxfile(path):
    file=load_workbook(path)
    sheets=file.sheetnames
    sheet=file.worksheets[0]
    # print(sheet.max_row)
    # print(sheet.max_column)
    # print(sheet.title)

    for lineNum in range(1,sheet.max_row+1):
        linelist = []
        for columnNum in  range(1,sheet.max_column+1):
             value=sheet.cell(row=lineNum,column=columnNum).value
             linelist.append(value)
        print(linelist)



path="D:\\pythonitem\\demo\\Demo7\\1.xlsx"
readxlsxfile(path)
